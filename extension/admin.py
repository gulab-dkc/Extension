from django.contrib import admin
from django.core.mail import send_mail
from .models import Extension,Employee
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.utils import timezone
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment,Border, Side
from openpyxl.utils import get_column_letter

admin.site.register(Employee)

@admin.register(Extension)
class ExtensionAdmin(admin.ModelAdmin):
    list_display = ("emp", "extension", "status", "created_at", "updated_at")

    def generate_excel(self):
        """
        Generate an Excel file containing usernames and extensions.
        - 14 records per column
        - Maximum of 3 columns
        - Extension and username in one cell, separate lines
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Intercom List"

        # Define thick black border
        thick_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Add heading with thick border and bold font
        ws.merge_cells('A1:C1')
        ws["A1"] = "Intercom List"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(bold=True, size=14)

        ws.merge_cells('A2:C2')
        ws["A2"] = f"Updated Date: {timezone.now().strftime('%d %b %Y')}"
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A2"].font = Font(bold=True, size=12)

        # Apply thick black border for A1:C20
        for row in ws.iter_rows(min_row=0, max_row=20, min_col=0, max_col=3):
            for cell in row:
                cell.border = thick_border

        # Adjust width and alignment for only 3 columns
        for col_index in range(1, 4):  # Columns A to C (3 columns max)
            col_letter = get_column_letter(col_index)
            ws.column_dimensions[col_letter].width = 25  # Adjust width for better view
            for row in range(1, ws.max_row + 1):
                ws[f"{col_letter}{row}"].alignment = Alignment(
                    wrap_text=True, vertical="center", horizontal="center"
                )

        # Add data: 14 records per column, max 3 columns
        row_index, col_index = 3, 1
        for idx, extension in enumerate(Extension.objects.all(), start=1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.value = f"{extension.extension}\n{extension.emp.name}"
            cell.font = Font(bold=True,size=10)  # Make text bold
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )

            row_index += 1
            if idx % 14 == 0:
                row_index = 3
                col_index += 1
                if col_index > 3:
                    col_index = 1
                    row_index += 15

        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        subject = "Intercom Details Updated" if change else "New Intercom Details Created"

        # Generate Excel
        excel_file = self.generate_excel()

        # Send personalized email with Excel attachment
        users = Employee.objects.exclude(email="")  # Get all users with email
        for user in users:
            message = f"""
            Hello {user.name},<br><br>
            Please check the updated intercom list attached.<br><br>
            Regards,<br>
            DKC Exports
            """

            email = EmailMultiAlternatives(
                subject=subject,
                body="Please view this email in HTML format.",
                from_email=settings.EMAIL_HOST_USER,
                to=[user.email],
            )
            email.attach_alternative(message, "text/html")
            email.attach(
                f"Telecom_Report_{timezone.now().strftime('%d %b %Y')}.xlsx",
                excel_file.read(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            email.send(fail_silently=False)
            excel_file.seek(0)




